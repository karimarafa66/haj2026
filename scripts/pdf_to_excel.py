#!/usr/bin/env python3
"""
pdf_to_excel.py  —  Extract Haj pilgrim data from PDF and write an Excel file

Usage:
    python pdf_to_excel.py                        # uses data-sources/hajPDF.pdf
    python pdf_to_excel.py path/to/file.pdf       # custom PDF → excel-data/
    python pdf_to_excel.py path/to/file.pdf out.xlsx  # custom PDF + output name

Requirements:
    pip install pdfplumber openpyxl
"""

import sys
import io
import re
import unicodedata
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR  = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
PDF_PATH    = Path(sys.argv[1]) if len(sys.argv) > 1 else PROJECT_DIR / 'data-sources' / 'hajPDF.pdf'

if len(sys.argv) > 2:
    OUT_PATH = Path(sys.argv[2])
elif PDF_PATH == (PROJECT_DIR / 'data-sources' / 'hajPDF.pdf'):
    pdf_name = PDF_PATH.stem
    OUT_PATH = PROJECT_DIR / 'excel-data' / f'{pdf_name}_new.xlsx'
else:
    OUT_PATH = PDF_PATH.with_suffix('.xlsx')

if not PDF_PATH.exists():
    print(f'ERROR: PDF not found: {PDF_PATH}')
    sys.exit(1)

try:
    import pdfplumber
except ImportError:
    print('pdfplumber not installed. Run: pip install pdfplumber')
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print('openpyxl not installed. Run: pip install openpyxl')
    sys.exit(1)

# ── reuse all parsing logic from pdf_to_data.py ──────────────────────────────
# Rather than duplicating, import the parse functions directly.
sys.path.insert(0, str(SCRIPT_DIR))
try:
    from pdf_to_data import parse_pdf, post_process
except ImportError:
    print('ERROR: pdf_to_data.py not found in the same folder.')
    sys.exit(1)


# ── Column definitions (Arabic headers, right-to-left order) ─────────────────
COLUMNS = [
    ('م',            'row_num'),
    ('رقم الطلب',    'request_num'),
    ('الرقم القومي', 'national_id'),
    ('رقم الجواز',   'passport'),
    ('الاسم',        'name'),
    ('الغرفة',       'room'),
    ('الدور',        'floor'),
    ('الجهة',        'region'),
    ('كود الرحلة',   'flight_code'),
    ('صلة القرابة',  'relation'),
]

# ── Styling constants ─────────────────────────────────────────────────────────
COLOR_HEADER_BG  = '1B4332'   # dark green
COLOR_HEADER_FG  = 'F7F0E0'   # cream
COLOR_ALT_ROW    = 'EAF4EE'   # light green
COLOR_WHITE      = 'FFFFFF'
COLOR_BORDER     = 'A8D5B5'
COLOR_FLOOR_BG   = '2D6A4F'   # medium green
COLOR_FLOOR_FG   = 'FFFFFF'

FONT_ARABIC  = 'Cairo'
FONT_LATIN   = 'Calibri'

RTL = Alignment(horizontal='right', vertical='center', wrap_text=False, readingOrder=2)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=False, readingOrder=2)
LTR = Alignment(horizontal='left',  vertical='center', wrap_text=False, readingOrder=1)

def thin_border(color=COLOR_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill():
    return PatternFill('solid', fgColor=COLOR_HEADER_BG)

def alt_fill(i):
    c = COLOR_ALT_ROW if i % 2 == 0 else COLOR_WHITE
    return PatternFill('solid', fgColor=c)

def floor_fill():
    return PatternFill('solid', fgColor=COLOR_FLOOR_BG)


# ── Build workbook ────────────────────────────────────────────────────────────
def build_excel(records: list[dict], out_path: Path):
    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = 'ملخص'
    ws_summary.sheet_view.rightToLeft = True

    from collections import Counter
    floors  = Counter(r['floor']  for r in records)
    regions = Counter(r['region'] for r in records)
    rooms   = len(set(r['room']   for r in records))

    summary_data = [
        ('إجمالي الحجاج',  len(records)),
        ('عدد الغرف',       rooms),
        ('عدد الأدوار',     len(floors)),
        ('عدد المحافظات',   len(regions)),
    ]

    ws_summary.column_dimensions['A'].width = 22
    ws_summary.column_dimensions['B'].width = 16

    # Title
    ws_summary.merge_cells('A1:B1')
    title_cell = ws_summary['A1']
    title_cell.value = 'ملخص إحصائي — تقرير الحجاج'
    title_cell.font      = Font(name=FONT_ARABIC, bold=True, size=14, color=COLOR_HEADER_FG)
    title_cell.fill      = header_fill()
    title_cell.alignment = CTR
    title_cell.border    = thin_border()
    ws_summary.row_dimensions[1].height = 28

    for i, (label, val) in enumerate(summary_data, start=2):
        ca, cb = ws_summary.cell(i, 1), ws_summary.cell(i, 2)
        ca.value = label
        cb.value = val
        for c in (ca, cb):
            c.fill      = alt_fill(i)
            c.border    = thin_border()
            c.alignment = CTR
            c.font      = Font(name=FONT_ARABIC, size=11, bold=(c == ca))
        ws_summary.row_dimensions[i].height = 22

    # Floor breakdown
    row = len(summary_data) + 3
    ws_summary.cell(row, 1).value = 'الدور'
    ws_summary.cell(row, 2).value = 'عدد الحجاج'
    for col in (1, 2):
        c = ws_summary.cell(row, col)
        c.font = Font(name=FONT_ARABIC, bold=True, color=COLOR_HEADER_FG, size=11)
        c.fill = floor_fill()
        c.alignment = CTR
        c.border = thin_border()
    ws_summary.row_dimensions[row].height = 22
    row += 1

    floor_order = [
        'الدور الأول','الدور الثاني','الدور الثالث','الدور الرابع','الدور الخامس',
        'الدور السادس','الدور السابع','الدور الثامن','الدور التاسع','الدور العاشر',
        'الدور الحادي عشر','الدور الثاني عشر','الدور الثالث عشر',
    ]
    sorted_floors = sorted(floors.items(), key=lambda x: floor_order.index(x[0]) if x[0] in floor_order else 99)
    for i, (floor, cnt) in enumerate(sorted_floors):
        ca, cb = ws_summary.cell(row, 1), ws_summary.cell(row, 2)
        ca.value, cb.value = floor, cnt
        for c in (ca, cb):
            c.fill = alt_fill(i)
            c.border = thin_border()
            c.alignment = CTR
            c.font = Font(name=FONT_ARABIC, size=10)
        ws_summary.row_dimensions[row].height = 20
        row += 1

    # Region breakdown
    row += 1
    ws_summary.cell(row, 1).value = 'المحافظة'
    ws_summary.cell(row, 2).value = 'عدد الحجاج'
    for col in (1, 2):
        c = ws_summary.cell(row, col)
        c.font = Font(name=FONT_ARABIC, bold=True, color=COLOR_HEADER_FG, size=11)
        c.fill = floor_fill()
        c.alignment = CTR
        c.border = thin_border()
    ws_summary.row_dimensions[row].height = 22
    row += 1

    for i, (region, cnt) in enumerate(sorted(regions.items(), key=lambda x: -x[1])):
        ca, cb = ws_summary.cell(row, 1), ws_summary.cell(row, 2)
        ca.value, cb.value = region or '—', cnt
        for c in (ca, cb):
            c.fill = alt_fill(i)
            c.border = thin_border()
            c.alignment = CTR
            c.font = Font(name=FONT_ARABIC, size=10)
        ws_summary.row_dimensions[row].height = 20
        row += 1

    # ── All pilgrims sheet ────────────────────────────────────────────────────
    ws_all = wb.create_sheet('جميع الحجاج')
    ws_all.sheet_view.rightToLeft = True
    _write_data_sheet(ws_all, records, 'جميع الحجاج')

    # ── One sheet per floor ───────────────────────────────────────────────────
    for floor_name in floor_order:
        floor_records = [r for r in records if r['floor'] == floor_name]
        if not floor_records:
            continue
        # Short sheet name (Excel limit: 31 chars)
        sheet_name = floor_name.replace('الدور ', '')
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.rightToLeft = True
        _write_data_sheet(ws, floor_records, floor_name)

    wb.save(str(out_path))
    print(f'Written: {out_path}  ({len(records)} records, {len(wb.sheetnames)} sheets)')


def _write_data_sheet(ws, records: list[dict], title: str):
    headers = [col[0] for col in COLUMNS]
    fields  = [col[1] for col in COLUMNS]

    # Column widths
    widths = {
        'row_num': 6, 'request_num': 12, 'national_id': 18,
        'passport': 14, 'name': 34, 'room': 9, 'floor': 18,
        'region': 16, 'flight_code': 12, 'relation': 16,
    }
    for col_i, field in enumerate(fields, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = widths.get(field, 14)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    tc = ws.cell(1, 1)
    tc.value     = f'تقرير الحجاج — {title}  |  إجمالي: {len(records)} حاج'
    tc.font      = Font(name=FONT_ARABIC, bold=True, size=13, color=COLOR_HEADER_FG)
    tc.fill      = header_fill()
    tc.alignment = CTR
    tc.border    = thin_border()
    ws.row_dimensions[1].height = 26

    # Header row
    for col_i, header in enumerate(headers, start=1):
        c = ws.cell(2, col_i)
        c.value     = header
        c.font      = Font(name=FONT_ARABIC, bold=True, size=11, color=COLOR_HEADER_FG)
        c.fill      = floor_fill()
        c.alignment = CTR
        c.border    = thin_border()
    ws.row_dimensions[2].height = 22

    # Freeze header rows
    ws.freeze_panes = 'A3'

    # Data rows
    for row_i, rec in enumerate(records, start=3):
        fill = alt_fill(row_i)
        for col_i, field in enumerate(fields, start=1):
            c   = ws.cell(row_i, col_i)
            val = rec.get(field, '')

            # national_id and passport stay LTR
            if field in ('national_id', 'passport'):
                c.alignment = LTR
                c.font      = Font(name=FONT_LATIN, size=10)
            elif field in ('row_num', 'request_num', 'room', 'flight_code'):
                c.alignment = CTR
                c.font      = Font(name=FONT_ARABIC, size=10)
            else:
                c.alignment = RTL
                c.font      = Font(name=FONT_ARABIC, size=10)

            c.value  = val if val else '—'
            c.fill   = fill
            c.border = thin_border()
        ws.row_dimensions[row_i].height = 18

    # Auto-filter on header row
    ws.auto_filter.ref = f'A2:{get_column_letter(len(COLUMNS))}{len(records) + 2}'


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f'Parsing: {PDF_PATH}')
    records = parse_pdf(PDF_PATH)
    print(f'Raw records: {len(records)}')
    records = post_process(records)
    print(f'After cleanup: {len(records)}')
    build_excel(records, OUT_PATH)
    print('\nDone. Open the .xlsx file in Excel.')


if __name__ == '__main__':
    main()
